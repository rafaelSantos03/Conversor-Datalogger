from flask import Flask, request, render_template, redirect, url_for, send_file
import pandas as pd
import os
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import io
import webbrowser
import time
from threading import Thread
import os
base_path = os.path.dirname(os.path.abspath(__file__))
logo_path = os.path.join(base_path, 'static', 'images', 'logo.png')


# Configurações do Flask
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

# Criar pasta para uploads, se não existir
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

@app.route('/')
def index():
    """Página inicial que exibe o formulário de upload"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Recebe o arquivo enviado pelo usuário e processa"""
    if 'file' not in request.files:
        return "Erro: Nenhum arquivo enviado", 400

    file = request.files['file']

    if file.filename == '':
        return "Erro: Nenhum arquivo selecionado", 400

    if file:
        filename = file.filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Verificar qual modo de leitura usar
        reading_mode = detect_reading_mode(filepath)
        return redirect(url_for('resultado', filename=filename, mode=reading_mode))

def detect_reading_mode(filepath):
    """Detecta automaticamente qual modo de leitura usar baseado nas colunas do arquivo"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Verificar formato específico procurando pela tabela de dados
        for row_idx in range(1, min(50, ws.max_row + 1)):  # Procurar nas primeiras 50 linhas
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            # Verificar se é o formato específico com a tabela de dados
            has_datetime = any('data/hora' in val or ('data' in val and 'hora' in val) for val in row_values)
            has_temp = any('temperatura' in val and '°c' in val for val in row_values)
            has_humid = any('umidade' in val and ('hr' in val or '%hr' in val) for val in row_values)
            
            # Se encontrou as 3 colunas principais, é o formato específico
            if has_datetime and has_temp and has_humid:
                wb.close()
                print(f"Formato específico detectado na linha {row_idx}")
                return 'specific_format'
        
        # Procurar pelo formato de relatório nas primeiras 50 linhas
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            # Verificar se é o formato de relatório (N°., Temp, UR, Tempo)
            if ('n°.' in row_values or 'nº.' in row_values or 'n°' in row_values) and \
               ('temp' in row_values) and \
               ('ur' in row_values) and \
               ('tempo' in row_values):
                wb.close()
                print(f"Modo relatório detectado na linha {row_idx}")
                return 'report_mode'
        
        # Se chegou até aqui, usar modo atual como padrão
        print("Usando modo original como padrão")
        return 'current_mode'
        
    except Exception as e:
        print(f"Erro na detecção: {e}")
        return 'current_mode'

def process_report_mode(filepath, filename):
    """Processa o arquivo usando o modo de relatório"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Encontrar a linha dos cabeçalhos da tabela de dados
        header_row = None
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            # Procurar pela linha com N°., Temp, UR, Tempo
            if ('n°.' in row_values or 'nº.' in row_values or 'n°' in row_values) and \
               ('temp' in row_values) and \
               ('ur' in row_values) and \
               ('tempo' in row_values):
                header_row = row_idx
                break
        
        if header_row is None:
            raise Exception("Não foi possível encontrar os cabeçalhos da tabela de dados")
        
        wb.close()
        
        # Ler o arquivo Excel a partir da linha dos cabeçalhos
        skip_rows = header_row - 1
        df = pd.read_excel(filepath, skiprows=skip_rows)
        
        if df.empty:
            raise Exception("Arquivo vazio após aplicar skiprows")
        
        # Normalizar as colunas para minúsculas
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        print(f"Processando modo relatório com skiprows={skip_rows}, colunas: {list(df.columns)}")
        
        # Mapear as colunas (ignorando N°.)
        temp_col = None
        ur_col = None
        tempo_col = None
        
        for col in df.columns:
            if 'temp' in col and not 'tempo' in col:
                temp_col = col
            elif 'ur' in col:
                ur_col = col
            elif 'tempo' in col:
                tempo_col = col
        
        if not all([temp_col, ur_col, tempo_col]):
            raise Exception(f"Colunas necessárias não encontradas. Encontradas: {list(df.columns)}")
        
        # Filtrar apenas as colunas necessárias (ignorando N°.)
        df_filtered = df[[tempo_col, temp_col, ur_col]].copy()
        
        # Renomear para padrão
        df_filtered.columns = ['tempo', 'temperatura', 'umidade']
        
        # Converter a coluna 'tempo' para datetime
        df_filtered['tempo'] = pd.to_datetime(df_filtered['tempo'], errors='coerce')
        
        # CORREÇÃO: Converter colunas de temperatura e umidade para numérico
        # Remover caracteres não numéricos e converter para float
        df_filtered['temperatura'] = pd.to_numeric(
            df_filtered['temperatura'].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
            errors='coerce'
        )
        
        df_filtered['umidade'] = pd.to_numeric(
            df_filtered['umidade'].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
            errors='coerce'
        )
        
        # Remover linhas com dados inválidos
        df_filtered = df_filtered.dropna(subset=['tempo', 'temperatura', 'umidade'])
        
        if df_filtered.empty:
            raise Exception("Nenhum dado válido encontrado após conversão")
        
        # Extrair apenas a data (ignorando a hora)
        df_filtered['data'] = df_filtered['tempo'].dt.date
        
        # Calcular os valores máximos e mínimos por dia
        result = df_filtered.groupby('data').agg(
            Temp_Max=('temperatura', 'max'),
            Temp_Min=('temperatura', 'min'),
            Umid_Max=('umidade', 'max'),
            Umid_Min=('umidade', 'min')
        ).reset_index()
        
        # Formatar a coluna 'data' no formato de data brasileiro (DD/MM/YYYY)
        result['data'] = result['data'].apply(lambda x: x.strftime('%d/%m/%Y'))
        
        # Ajustar a formatação dos números
        result['Temp_Max'] = result['Temp_Max'].round(2)
        result['Temp_Min'] = result['Temp_Min'].round(2)
        result['Umid_Max'] = result['Umid_Max'].round(2)
        result['Umid_Min'] = result['Umid_Min'].round(2)
        
        # Renomear as colunas para exibir os nomes personalizados
        result = result.rename(columns={
            'data': 'Data',
            'Temp_Max': 'Temperatura Máxima (°C)',
            'Temp_Min': 'Temperatura Mínima (°C)',
            'Umid_Max': 'Umidade Máxima (%)',
            'Umid_Min': 'Umidade Mínima (%)'
        })
        
        # Salvar o dataframe como variável global para uso na geração do PDF
        global latest_result
        latest_result = result
        
        # Converter a tabela para HTML e remover quebras de linha extras
        table_html = result.to_html(classes='table table-striped table-bordered', index=False)
        table_html = table_html.replace("\n", "")
        
        # Exibir resultados
        return render_template(
            'resultado.html',
            table=table_html,
            filename=filename
        )
        
    except Exception as e:
        print(f"Erro no processamento do modo relatório: {e}")
        return (
            f"Erro: Não foi possível processar o arquivo. {str(e)}",
            400,
        )

@app.route('/resultado/<filename>', methods=['GET', 'POST'])
def resultado(filename):
    """Processa o arquivo Excel e exibe os resultados filtrados"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    reading_mode = request.args.get('mode', 'current_mode')
    
    try:
        if reading_mode == 'specific_format':
            return process_specific_format(filepath, filename)
        elif reading_mode == 'report_mode':
            return process_report_mode(filepath, filename)
        elif reading_mode == 'new_mode':
            return process_new_mode(filepath, filename)
        else:
            return process_current_mode(filepath, filename)
    except Exception as e:
        return f"Erro ao processar o arquivo: {e}", 500

def process_current_mode(filepath, filename):
    """Processa o arquivo usando o modo de leitura atual (original)"""
    # Tentar diferentes configurações de skiprows
    for skip_rows in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        try:
            df = pd.read_excel(filepath, skiprows=skip_rows)
            if df.empty or len(df.columns) < 2:
                continue
                
            # Normalizar as colunas para minúsculas e remover espaços
            df.columns = df.columns.astype(str).str.strip().str.lower().str.replace(' ', '_')
            
            print(f"Tentando modo original com skiprows={skip_rows}, colunas: {list(df.columns)}")
            
            # Buscar colunas de forma mais flexível
            date_col = None
            temp_col = None
            humid_col = None
            time_col = None
            
            # Procurar coluna de data/tempo
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['data', 'date', 'tempo', 'time', 'hora']):
                    if 'temp' not in col_lower:  # Evitar confundir com temperatura
                        if date_col is None:
                            date_col = col
                        elif 'time' in col_lower or 'hora' in col_lower:
                            time_col = col
            
            # Procurar coluna de temperatura
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['temp', 'temperatura', 'oc', '°c']):
                    temp_col = col
                    break
            
            # Procurar coluna de umidade
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['umid', 'humid', 'rh', '%rh', 'hr', '%hr']):
                    humid_col = col
                    break
            
            # Verificar se encontrou as colunas essenciais
            if not all([date_col, temp_col, humid_col]):
                print(f"Colunas não encontradas - Data: {date_col}, Temp: {temp_col}, Umid: {humid_col}")
                continue
            
            print(f"Colunas identificadas - Data: {date_col}, Temp: {temp_col}, Umid: {humid_col}, Time: {time_col}")
            
            # Criar coluna datetime
            if time_col and time_col in df.columns:
                # Combinar data e hora se houver coluna de tempo separada
                df['datetime'] = pd.to_datetime(
                    df[date_col].astype(str) + ' ' + df[time_col].astype(str),
                    errors='coerce'
                )
            else:
                # Usar apenas a coluna de data
                df['datetime'] = pd.to_datetime(df[date_col], errors='coerce')
            
            # Remover linhas com datas inválidas
            df = df.dropna(subset=['datetime'])
            
            if df.empty:
                print("Nenhuma data válida encontrada")
                continue
            
            # Converter temperatura e umidade para numérico
            df['temperatura_num'] = pd.to_numeric(
                df[temp_col].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
                errors='coerce'
            )
            
            df['umidade_num'] = pd.to_numeric(
                df[humid_col].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
                errors='coerce'
            )
            
            # Remover linhas com valores inválidos
            df = df.dropna(subset=['temperatura_num', 'umidade_num'])
            
            if df.empty:
                print("Nenhum dado numérico válido encontrado")
                continue
            
            # Extrair apenas a data (ignorando a hora)
            df['date'] = df['datetime'].dt.date
            
            # Calcular os valores máximos e mínimos por dia
            result = df.groupby('date').agg(
                Temp_Max=('temperatura_num', 'max'),
                Temp_Min=('temperatura_num', 'min'),
                Umid_Max=('umidade_num', 'max'),
                Umid_Min=('umidade_num', 'min')
            ).reset_index()
            
            # Formatar a coluna 'date' no formato de data brasileiro (DD/MM/YYYY)
            result['date'] = result['date'].apply(lambda x: x.strftime('%d/%m/%Y'))
            
            # Ajustar a formatação dos números
            result['Temp_Max'] = result['Temp_Max'].round(2)
            result['Temp_Min'] = result['Temp_Min'].round(2)
            result['Umid_Max'] = result['Umid_Max'].round(2)
            result['Umid_Min'] = result['Umid_Min'].round(2)
            
            # Renomear as colunas para exibir os nomes personalizados
            result = result.rename(columns={
                'date': 'Data',
                'Temp_Max': 'Temperatura Máxima (°C)',
                'Temp_Min': 'Temperatura Mínima (°C)',
                'Umid_Max': 'Umidade Máxima (%)',
                'Umid_Min': 'Umidade Mínima (%)'
            })
            
            # Salvar o dataframe como variável global para uso na geração do PDF
            global latest_result
            latest_result = result
            
            # Converter a tabela para HTML e remover quebras de linha extras
            table_html = result.to_html(classes='table table-striped table-bordered', index=False)
            table_html = table_html.replace("\n", "")
            
            # Exibir resultados
            return render_template(
                'resultado.html',
                table=table_html,
                filename=filename
            )
                    
        except Exception as e:
            print(f"Erro com skiprows={skip_rows}: {e}")
            continue
    
    # Se chegou até aqui, não conseguiu processar
    return (
        f"Erro: Não foi possível identificar as colunas necessárias no arquivo. "
        f"Verifique se o arquivo contém colunas de data/tempo, temperatura e umidade. "
        f"Formatos suportados: Data, Temperatura (°C), Umidade (%), etc.",
        400,
    )

def process_new_mode(filepath, filename):
    """Processa o arquivo usando o novo modo de leitura"""
    try:
        # Encontrar a linha correta dos cabeçalhos
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        header_row = None
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            if ('id' in row_values and 
                any('data' in val and 'hora' in val for val in row_values) and
                any('temperatura' in val for val in row_values) and
                any('umidade' in val for val in row_values)):
                header_row = row_idx
                break
        
        wb.close()
        
        if header_row is None:
            raise Exception("Não foi possível encontrar os cabeçalhos corretos")
        
        # Ler o arquivo com o skiprows correto
        skip_rows = header_row - 1
        df = pd.read_excel(filepath, skiprows=skip_rows)
        
        if df.empty:
            raise Exception("Arquivo vazio após aplicar skiprows")
            
        # Normalizar as colunas para minúsculas
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        print(f"Processando novo modo com skiprows={skip_rows}, colunas: {list(df.columns)}")
        
        # Mapear as colunas encontradas
        id_col = None
        datetime_col = None
        temp_col = None
        humid_col = None
        
        for col in df.columns:
            if col == 'id':
                id_col = col
            elif 'data' in col and 'hora' in col:
                datetime_col = col
            elif 'temperatura' in col:
                temp_col = col
            elif 'umidade' in col:
                humid_col = col
        
        if not all([datetime_col, temp_col, humid_col]):
            raise Exception(f"Colunas necessárias não encontradas. Encontradas: {list(df.columns)}")
        
        # Filtrar as colunas necessárias (ignorando o campo 'id')
        df_filtered = df[[datetime_col, temp_col, humid_col]].copy()
        
        # Renomear para padrão
        df_filtered.columns = ['data/hora', 'temperatura', 'umidade']
        
        # Converter a coluna 'data/hora' para datetime e extrair apenas a data
        df_filtered['data/hora'] = pd.to_datetime(df_filtered['data/hora'], errors='coerce')
        
        # Remover linhas com datas inválidas
        df_filtered = df_filtered.dropna(subset=['data/hora'])
        
        if df_filtered.empty:
            raise Exception("Nenhuma data válida encontrada")
        
        # Extrair apenas a data (ignorando a hora)
        df_filtered['data'] = df_filtered['data/hora'].dt.date
        
        # Calcular os valores máximos e mínimos por dia
        result = df_filtered.groupby('data').agg(
            Temp_Max=('temperatura', 'max'),
            Temp_Min=('temperatura', 'min'),
            Umid_Max=('umidade', 'max'),
            Umid_Min=('umidade', 'min')
        ).reset_index()
        
        # Formatar a coluna 'data' no formato de data brasileiro (DD/MM/YYYY)
        result['data'] = result['data'].apply(lambda x: x.strftime('%d/%m/%Y'))
        
        # Ajustar a formatação dos números
        result['Temp_Max'] = result['Temp_Max'].round(2)
        result['Temp_Min'] = result['Temp_Min'].round(2)
        result['Umid_Max'] = result['Umid_Max'].round(2)
        result['Umid_Min'] = result['Umid_Min'].round(2)
        
        # Renomear as colunas para exibir os nomes personalizados
        result = result.rename(columns={
            'data': 'Data',
            'Temp_Max': 'Temperatura Máxima (°C)',
            'Temp_Min': 'Temperatura Mínima (°C)',
            'Umid_Max': 'Umidade Máxima (%)',
            'Umid_Min': 'Umidade Mínima (%)'
        })
        
        # Salvar o dataframe como variável global para uso na geração do PDF
        global latest_result
        latest_result = result
        
        # Converter a tabela para HTML e remover quebras de linha extras
        table_html = result.to_html(classes='table table-striped table-bordered', index=False)
        table_html = table_html.replace("\n", "")
        
        # Exibir resultados
        return render_template(
            'resultado.html',
            table=table_html,
            filename=filename
        )
        
    except Exception as e:
        print(f"Erro no processamento do novo modo: {e}")
        return (
            f"Erro: Não foi possível processar o arquivo. {str(e)}",
            400,
        )

@app.route('/gerar_pdf', methods=['GET', 'POST'])
def gerar_pdf():
    """Gera um PDF com os resultados filtrados"""
    try:
        # Receber os parâmetros do formulário
        param1 = request.form.get('param1')  # Formulação
        param2 = request.form.get('param2')  # Revisão
        param3 = request.form.get('param3')  # 'aprovado' ou 'reprovado'
        param4 = request.form.get('param4')  # Data fornecida no formulário
        param5 = request.form.get('param5')  # Número do estudo
        param6 = request.form.get('param6')  # Código do equipamento
        param7 = request.form.get('param7')  # Número do ensaio
        param8 = request.form.get('param8')  # Local de leitura do equipamento

        # Para voltar a ser dinamico remova os 'disable' dos html e apague da linha 129 a 132
        param1 = "FOR.2.031"  # Formulação
        param2 = "Rev. 00"  # Revisão
        param3 = "Aprovado"  # 'aprovado' ou 'reprovado'

        # Garantir que os valores são strings
        param1 = str(param1) if param1 else "oi1"
        param2 = str(param2) if param2 else "oi2"
        param3 = str(param3) if param3 else "oi3"
        param4 = str(param4) if param4 else "24/03/2025"
        param5 = str(param5) if param5 else "oi5"
        param6 = str(param6) if param6 else "oi6"
        param7 = str(param7) if param7 else "oi7"
        param8 = str(param8) if param8 else "oi8"

        # Verifica se os resultados estão disponíveis
        if 'latest_result' not in globals() or latest_result.empty:
            return "Erro: Nenhum dado disponível para gerar o PDF.", 400

        # Configura o PDF
        buffer = io.BytesIO()

        # Calcular o número total de páginas
        items_per_page = 30  # Defina o número de itens por página
        total_items = len(latest_result)  # Número total de linhas (excluindo cabeçalhos)
        total_pages = (total_items // items_per_page) + (1 if total_items % items_per_page > 0 else 0)


        def add_header(canvas, doc, is_first_page, total_pages, param1, param2, param3, param4, param5, param6, param7, param8):
            """Função para adicionar o cabeçalho"""
            canvas.saveState()

            # Cabeçalho principal
            canvas.rect(9, 720, 592, 60)
            import sys
            import os

            def get_resource_path(relative_path):
                """Retorna o caminho correto do recurso, seja no ambiente normal ou no .exe."""
                if getattr(sys, 'frozen', False):  # Se estiver rodando no PyInstaller (.exe)
                    base_path = sys._MEIPASS
                else:
                   if getattr(sys, 'frozen', False):  # Se estiver rodando num executável (.exe)
                      base_path = sys._MEIPASS
                   else:
                    base_path = os.path.dirname(os.path.abspath(sys.argv[0]))  # Usa o primeiro argumento ao invés de __file__
                
                return os.path.join(base_path, relative_path)

            logo_path = get_resource_path("static/images/logo.png")


            try:
                canvas.drawImage(logo_path, 35, 730, width=100, height=40, mask='auto')
            except Exception as e:
                raise Exception(f"Erro ao carregar a logo: {e}")
            canvas.line(150, 780, 150, 720)
            canvas.line(450, 780, 450, 720)
            canvas.line(450, 750, 601, 750) #linha horizontal do meio canto direito
            canvas.line(523, 780, 523, 720) #linha vertical do meio canto direito

            # Título e informações
            canvas.setFont("Helvetica-Bold", 13)
            canvas.drawString(160, 745, "DADOS DE TEMPERATURA E/OU UMIDADE")
            canvas.setFont("Helvetica", 10)
            canvas.drawString(455, 760, f"{param1}")  # Formulação
            canvas.setFont("Helvetica", 10)
            canvas.drawString(550, 760, f"{doc.page} / {total_pages}")  # Paginas
            canvas.drawString(470, 730, f"{param2}")  # Revisao
            canvas.drawString(540, 737, f"{param3}")  # 'aprovado' ou 'reprovado'
            canvas.drawString(535, 725, f"{param4}")  # Data

            # Tabela adicional (agora em todas as páginas)
            data = [["Número do estudo:", str(param5), "Código do equipamento:", str(param6)],
                    ["Número do ensaio:", str(param7), "Local de leitura do equipamento:", str(param8)]]
            table = Table(data, colWidths=[100,169,155,169], rowHeights=20)
            table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ]))

            table.wrapOn(canvas, 9, 670)
            table.drawOn(canvas, 9, 670)

            canvas.restoreState()

        def add_footer(canvas, doc):
            """Função para adicionar o rodapé com o campo de assinatura"""
            canvas.saveState()
            canvas.setFont("Helvetica", 10)
            canvas.drawString(40, 50, "Rubrica: ________________________________________________")
            canvas.drawString(40, 30, "Data: ___________________________________________________")
            canvas.restoreState()

        elements = []

        # Dados para a tabela principal
        data = [latest_result.columns.tolist()]
        data += latest_result.values.tolist()

        # Configuração da tabela
        table = Table(data, colWidths=[102, 123, 123, 123, 123], repeatRows=1)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        table.setStyle(style)
        elements.append(table)

        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=126)
        doc.build(
            elements,
            onFirstPage=lambda c, d: (add_header(c, d, True, total_pages, param1, param2, param3, param4, param5, param6, param7, param8), add_footer(c, d)),
            onLaterPages=lambda c, d: (add_header(c, d, False, total_pages, param1, param2, param3, param4, param5, param6, param7, param8), add_footer(c, d))
        )

        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False #,
            #download_name='resultado_filtrado.pdf'
        )

    except Exception as e:
        return f"Erro ao gerar o PDF: {e}", 500

def run_app():
    """Executa o servidor Flask com configurações otimizadas"""
    try:
        app.run(
            debug=False,  # Desabilitar debug para produção
            use_reloader=False,
            host='127.0.0.1',
            port=5000,
            threaded=True  # Permitir múltiplas conexões
        )
    except Exception as e:
        print(f"Erro ao iniciar o servidor Flask: {e}")

def check_server_ready(host='127.0.0.1', port=5000, timeout=15):
    """Verifica se o servidor está pronto para receber conexões"""
    start_time = time.time()
    
    # Primeiro, aguardar a porta estar em uso (servidor iniciado)
    while time.time() - start_time < timeout:
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(1)
                result = s.connect_ex((host, port))
                if result == 0:  # Conexão bem-sucedida
                    # Aguardar um pouco mais para o Flask estar completamente pronto
                    time.sleep(2)
                    
                    # Tentar fazer uma requisição HTTP
                    try:
                        response = urllib.request.urlopen(f'http://{host}:{port}/', timeout=3)
                        if response.getcode() == 200:
                            return True
                    except Exception:
                        # Se a requisição HTTP falhar, mas a porta está aberta,
                        # provavelmente o servidor está funcionando
                        return True
        except Exception:
            pass
        
        time.sleep(0.5)
    
    return False

def is_port_available(host='127.0.0.1', port=5000):
    """Verifica se a porta está disponível"""
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(1)
            result = s.connect_ex((host, port))
            return result != 0  # Porta disponível se conexão falhar
    except Exception:
        return True

def process_specific_format(filepath, filename):
    """Processa o arquivo com formato específico: ignora cabeçalhos e processa apenas id, Data/Hora, Temperatura[°C], Umidade[%Hr]"""
    try:
        # Tentar diferentes valores de skiprows para encontrar os cabeçalhos
        for skip_rows in range(0, 30):  # Aumentei para 30 para cobrir mais linhas de cabeçalho
            try:
                df = pd.read_excel(filepath, skiprows=skip_rows)
                if df.empty or len(df.columns) < 3:
                    continue
                
                # Verificar se encontrou a linha com as colunas da tabela de dados
                columns_lower = [str(col).lower().strip() for col in df.columns]
                
                # Procurar especificamente pelas colunas: id, Data/Hora, Temperatura[°C], Umidade[%Hr]
                has_id = any('id' == col.strip() for col in columns_lower)
                has_datetime = any('data/hora' in col or ('data' in col and 'hora' in col) for col in columns_lower)
                has_temp = any('temperatura' in col and '°c' in col for col in columns_lower)
                has_humid = any('umidade' in col and ('%hr' in col or 'hr' in col) for col in columns_lower)
                
                # Deve ter pelo menos Data/Hora, Temperatura e Umidade (id é opcional)
                if not all([has_datetime, has_temp, has_humid]):
                    continue
                
                print(f"Tabela de dados encontrada com skiprows={skip_rows}")
                print(f"Colunas encontradas: {list(df.columns)}")
                
                # Mapear as colunas específicas
                datetime_col = None
                temp_col = None
                humid_col = None
                
                for col in df.columns:
                    col_str = str(col).lower().strip()
                    if 'data/hora' in col_str or ('data' in col_str and 'hora' in col_str):
                        datetime_col = col
                    elif 'temperatura' in col_str and '°c' in col_str:
                        temp_col = col
                    elif 'umidade' in col_str and ('hr' in col_str or '%hr' in col_str):
                        humid_col = col
                
                if not all([datetime_col, temp_col, humid_col]):
                    print(f"Colunas não mapeadas corretamente: DateTime={datetime_col}, Temp={temp_col}, Humid={humid_col}")
                    continue
                
                print(f"Colunas mapeadas - DateTime: {datetime_col}, Temp: {temp_col}, Umid: {humid_col}")
                
                # Filtrar apenas as colunas necessárias (IGNORAR id e outras)
                df_filtered = df[[datetime_col, temp_col, humid_col]].copy()
                
                # Renomear para facilitar o processamento
                df_filtered.columns = ['data_hora', 'temperatura', 'umidade']
                
                # Remover linhas vazias ou com dados inválidos
                df_filtered = df_filtered.dropna(how='all')
                
                # Converter a coluna data_hora para datetime
                df_filtered['data_hora'] = pd.to_datetime(df_filtered['data_hora'], errors='coerce')
                
                # Remover linhas com datas inválidas
                df_filtered = df_filtered.dropna(subset=['data_hora'])
                
                if df_filtered.empty:
                    print("Nenhuma data válida encontrada")
                    continue
                
                # Converter temperatura e umidade para numérico
                df_filtered['temperatura'] = pd.to_numeric(
                    df_filtered['temperatura'].astype(str).str.replace(r'[^0-9.,\\-]', '', regex=True).str.replace(',', '.'),
                    errors='coerce'
                )
                
                df_filtered['umidade'] = pd.to_numeric(
                    df_filtered['umidade'].astype(str).str.replace(r'[^0-9.,\\-]', '', regex=True).str.replace(',', '.'),
                    errors='coerce'
                )
                
                # Remover linhas com valores numéricos inválidos
                df_filtered = df_filtered.dropna(subset=['temperatura', 'umidade'])
                
                if df_filtered.empty:
                    print("Nenhum dado numérico válido encontrado")
                    continue
                
                print(f"Dados processados: {len(df_filtered)} registros válidos")
                
                # Extrair apenas a data (IGNORAR a hora conforme solicitado)
                df_filtered['data'] = df_filtered['data_hora'].dt.date
                
                # Aplicar a MESMA LÓGICA dos outros modos: máximas e mínimas por data
                result = df_filtered.groupby('data').agg(
                    Temp_Max=('temperatura', 'max'),
                    Temp_Min=('temperatura', 'min'),
                    Umid_Max=('umidade', 'max'),
                    Umid_Min=('umidade', 'min')
                ).reset_index()
                
                # Formatar a coluna 'data' no formato brasileiro (DD/MM/YYYY)
                result['data'] = result['data'].apply(lambda x: x.strftime('%d/%m/%Y'))
                
                # Ajustar a formatação dos números (2 casas decimais)
                result['Temp_Max'] = result['Temp_Max'].round(2)
                result['Temp_Min'] = result['Temp_Min'].round(2)
                result['Umid_Max'] = result['Umid_Max'].round(2)
                result['Umid_Min'] = result['Umid_Min'].round(2)
                
                # Renomear as colunas para exibição (MESMO PADRÃO dos outros modos)
                result = result.rename(columns={
                    'data': 'Data',
                    'Temp_Max': 'Temperatura Máxima (°C)',
                    'Temp_Min': 'Temperatura Mínima (°C)',
                    'Umid_Max': 'Umidade Máxima (%)',
                    'Umid_Min': 'Umidade Mínima (%)'
                })
                
                print(f"Resultado final: {len(result)} dias processados")
                
                # Salvar resultado globalmente para geração do PDF
                global latest_result
                latest_result = result
                
                # Converter para HTML
                table_html = result.to_html(classes='table table-striped table-bordered', index=False)
                table_html = table_html.replace("\n", "")
                
                return render_template(
                    'resultado.html',
                    table=table_html,
                    filename=filename
                )
                
            except Exception as e:
                print(f"Erro com skiprows={skip_rows}: {e}")
                continue
        
        # Se não conseguiu processar
        raise Exception("Não foi possível encontrar a tabela de dados no arquivo")
        
    except Exception as e:
        print(f"Erro no processamento do formato específico: {e}")
        return (
            f"Erro: Não foi possível processar o arquivo no formato específico. {str(e)}",
            400,
        )

if __name__ == '__main__':
    print("=== Conversor Datalogger ===")
    print("Iniciando aplicação...")
    
    # Verificar se a porta está disponível
    if not is_port_available():
        print("Aviso: A porta 5000 já está em uso. Tentando continuar...")
    
    # Iniciar o servidor Flask em thread separada
    print("Iniciando servidor Flask...")
    server_thread = Thread(target=run_app)
    server_thread.daemon = True
    server_thread.start()
    
    # Aguardar o servidor estar pronto
    print("Aguardando servidor ficar pronto...")
    
    # Aguardar um tempo fixo para o servidor iniciar
    time.sleep(4)
    
    # Verificar se o servidor está respondendo
    server_ready = check_server_ready()
    
    if server_ready:
        print("✅ Servidor iniciado com sucesso!")
    else:
        print("⚠️ Servidor pode estar iniciando... Tentando abrir navegador mesmo assim.")
    
    print("🌐 Abrindo navegador...")
    
    # Abrir o navegador
    try:
        webbrowser.open("http://127.0.0.1:5000/")
        print("✅ Aplicação aberta no navegador!")
    except Exception as e:
        print(f"Erro ao abrir navegador: {e}")
        print("Acesse manualmente: http://127.0.0.1:5000/")
    
    print("\n📋 Instruções:")
    print("- A aplicação está rodando em: http://127.0.0.1:5000/")
    print("- Para encerrar, pressione Ctrl+C")
    print("- Se houver problemas, aguarde alguns segundos e recarregue a página")
    
    # Manter o programa rodando
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n🛑 Encerrando aplicação...")
        print("✅ Aplicação encerrada com sucesso!")